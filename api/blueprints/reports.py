"""
Reports & Analytics Blueprint
Provides intelligent metrics generation with background data refresh.
"""
from flask import Blueprint, request, jsonify
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
import threading
import time
from collections import defaultdict
from utils.decorators import handle_api_error, json_response, log_request as log_decorator, require_credentials
from utils.db import get_db, init_db
from core.api import get_api_client
logger = logging.getLogger(__name__)
reports_bp = Blueprint('reports', __name__)
# Background refresh state
_refresh_state = {
    'is_running': False,
    'last_refresh': None,
    'next_refresh': None,
    'status': 'idle',
    'progress': 0,
    'error': None
}
_refresh_lock = threading.Lock()
# ============================================================================
# Database Schema for Reports Cache
# ============================================================================
SCHEMA_REPORTS = """
CREATE TABLE IF NOT EXISTS reports_cache (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    report_type TEXT NOT NULL,
    service_desk_id TEXT,
    queue_id TEXT,
    data TEXT NOT NULL,
    generated_at TEXT NOT NULL,
    expires_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_reports_type ON reports_cache(report_type);
CREATE INDEX IF NOT EXISTS idx_reports_desk ON reports_cache(service_desk_id);
CREATE INDEX IF NOT EXISTS idx_reports_expires ON reports_cache(expires_at);
"""
SCHEMA_ML_ANALYSIS = """
CREATE TABLE IF NOT EXISTS ml_analysis_cache (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    service_desk_id TEXT NOT NULL,
    queue_id TEXT NOT NULL,
    data TEXT NOT NULL,
    generated_at TEXT NOT NULL,
    expires_at TEXT NOT NULL,
    UNIQUE(service_desk_id, queue_id)
);
CREATE INDEX IF NOT EXISTS idx_ml_desk ON ml_analysis_cache(service_desk_id);
CREATE INDEX IF NOT EXISTS idx_ml_queue ON ml_analysis_cache(queue_id);
CREATE INDEX IF NOT EXISTS idx_ml_expires ON ml_analysis_cache(expires_at);
"""
def init_reports_db():
    """Initialize reports database tables."""
    conn = get_db()
    try:
        for statement in SCHEMA_REPORTS.split(';'):
            if statement.strip():
                conn.execute(statement)
        for statement in SCHEMA_ML_ANALYSIS.split(';'):
            if statement.strip():
                conn.execute(statement)
        conn.commit()
        logger.info("✅ Reports & ML analysis database initialized")
    except Exception as e:
        logger.error(f"Failed to initialize reports/ML DB: {e}")
# Initialize on module load
init_reports_db()
# ============================================================================
# Intelligent Metrics Generator
# ============================================================================
class MetricsGenerator:
    """Generate intelligent metrics from issue data."""
    @staticmethod
    def analyze_issues(issues: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Generate comprehensive metrics from issues.
        Returns:
            - summary: Basic counts and percentages
            - trends: Time-based analysis
            - performance: SLA and response times
            - insights: AI-generated observations
        """
        if not issues:
            return {
                'summary': {'total': 0, 'open': 0, 'closed': 0},
                'trends': {},
                'performance': {},
                'insights': []
            }
        # Summary metrics
        total = len(issues)
        by_status = defaultdict(int)
        by_priority = defaultdict(int)
        by_assignee = defaultdict(int)
        by_type = defaultdict(int)
        created_dates = []
        resolved_dates = []
        response_times = []
        resolution_times = []
        for issue in issues:
            # Status distribution
            status = issue.get('status', {})
            status_name = status.get('name', 'Unknown') if isinstance(status, dict) else str(status)
            by_status[status_name] += 1
            # Priority distribution
            priority = issue.get('priority', {})
            priority_name = priority.get('name', 'Unknown') if isinstance(priority, dict) else str(priority)
            by_priority[priority_name] += 1
            # Assignee workload
            assignee = issue.get('assignee', {})
            assignee_name = assignee.get('displayName', 'Unassigned') if isinstance(assignee, dict) else 'Unassigned'
            by_assignee[assignee_name] += 1
            # Issue type
            issue_type = issue.get('issuetype', {})
            type_name = issue_type.get('name', 'Unknown') if isinstance(issue_type, dict) else str(issue_type)
            by_type[type_name] += 1
            # Time tracking
            created = issue.get('created')
            if created:
                try:
                    created_dates.append(datetime.fromisoformat(created.replace('Z', '+00:00')))
                except:
                    pass
            resolved = issue.get('resolutiondate')
            if resolved:
                try:
                    resolved_dates.append(datetime.fromisoformat(resolved.replace('Z', '+00:00')))
                except:
                    pass
        # Calculate summary
        summary = {
            'total': total,
            'open': sum(count for status, count in by_status.items() if status.lower() not in ['done', 'closed', 'resolved']),
            'closed': sum(count for status, count in by_status.items() if status.lower() in ['done', 'closed', 'resolved']),
            'by_status': dict(by_status),
            'by_priority': dict(by_priority),
            'by_assignee': dict(sorted(by_assignee.items(), key=lambda x: x[1], reverse=True)[:10]),  # Top 10
            'by_type': dict(by_type)
        }
        # Trends analysis
        trends = MetricsGenerator._analyze_trends(created_dates, resolved_dates)
        # Performance metrics
        performance = MetricsGenerator._analyze_performance(issues, created_dates, resolved_dates)
        # Generate insights
        insights = MetricsGenerator._generate_insights(summary, trends, performance)
        return {
            'summary': summary,
            'trends': trends,
            'performance': performance,
            'insights': insights,
            'generated_at': datetime.now().isoformat()
        }
    @staticmethod
    def _analyze_trends(created_dates: List[datetime], resolved_dates: List[datetime]) -> Dict[str, Any]:
        """Analyze time-based trends."""
        from datetime import timezone
        now = datetime.now(timezone.utc)
        # Last 7 days buckets
        last_7_days = []
        for i in range(7):
            date = now - timedelta(days=i)
            created_count = sum(1 for d in created_dates if d.date() == date.date())
            resolved_count = sum(1 for d in resolved_dates if d.date() == date.date())
            last_7_days.append({
                'date': date.strftime('%Y-%m-%d'),
                'created': created_count,
                'resolved': resolved_count
            })
        last_7_days.reverse()
        # Month-over-month (ensure timezone-aware comparison)
        from datetime import timezone
        current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
        # Safely compare with timezone-aware dates
        try:
            current_month_created = sum(1 for d in created_dates if d >= current_month_start)
            last_month_created = sum(1 for d in created_dates if last_month_start <= d < current_month_start)
        except TypeError:
            # Handle mixed timezone-aware/naive dates
            current_month_created = sum(1 for d in created_dates if d.replace(tzinfo=None) >= current_month_start.replace(tzinfo=None))
            last_month_created = sum(1 for d in created_dates if last_month_start.replace(tzinfo=None) <= d.replace(tzinfo=None) < current_month_start.replace(tzinfo=None))
        return {
            'last_7_days': last_7_days,
            'current_month_created': current_month_created,
            'last_month_created': last_month_created,
            'mom_growth': round((current_month_created - last_month_created) / last_month_created * 100, 1) if last_month_created > 0 else 0
        }
    @staticmethod
    def _analyze_performance(issues: List[Dict], created_dates: List[datetime], resolved_dates: List[datetime]) -> Dict[str, Any]:
        """Analyze performance metrics."""
        # Calculate average resolution time
        resolution_times = []
        for issue in issues:
            created = issue.get('created')
            resolved = issue.get('resolutiondate')
            if created and resolved:
                try:
                    created_dt = datetime.fromisoformat(created.replace('Z', '+00:00'))
                    resolved_dt = datetime.fromisoformat(resolved.replace('Z', '+00:00'))
                    delta = (resolved_dt - created_dt).total_seconds() / 3600  # hours
                    resolution_times.append(delta)
                except:
                    pass
        avg_resolution = round(sum(resolution_times) / len(resolution_times), 1) if resolution_times else 0
        # SLA compliance (placeholder)
        sla_compliant = sum(1 for issue in issues if issue.get('sla_status') == 'met')
        sla_total = len([i for i in issues if i.get('sla_status')])
        sla_rate = round(sla_compliant / sla_total * 100, 1) if sla_total > 0 else 0
        return {
            'avg_resolution_hours': avg_resolution,
            'median_resolution_hours': round(sorted(resolution_times)[len(resolution_times)//2], 1) if resolution_times else 0,
            'sla_compliance_rate': sla_rate,
            'total_resolved': len(resolved_dates),
            'resolution_times_distribution': {
                '< 1h': sum(1 for t in resolution_times if t < 1),
                '1-8h': sum(1 for t in resolution_times if 1 <= t < 8),
                '8-24h': sum(1 for t in resolution_times if 8 <= t < 24),
                '1-7d': sum(1 for t in resolution_times if 24 <= t < 168),
                '> 7d': sum(1 for t in resolution_times if t >= 168)
            }
        }
    @staticmethod
    def _generate_insights(summary: Dict, trends: Dict, performance: Dict) -> List[Dict[str, str]]:
        """Generate AI-like insights from metrics."""
        insights = []
        # Open vs closed ratio
        if summary['total'] > 0:
            open_rate = summary['open'] / summary['total'] * 100
            if open_rate > 70:
                insights.append({
                    'type': 'warning',
                    'title': 'High Open Issues',
                    'message': f"{open_rate:.0f}% of issues are still open. Consider reviewing backlog prioritization.",
                    'metric': f"{summary['open']} open / {summary['total']} total"
                })
            elif open_rate < 30:
                insights.append({
                    'type': 'success',
                    'title': 'Healthy Resolution Rate',
                    'message': f"Only {open_rate:.0f}% of issues remain open. Great job!",
                    'metric': f"{summary['closed']} resolved"
                })
        # Trend analysis
        mom_growth = trends.get('mom_growth', 0)
        if abs(mom_growth) > 20:
            insights.append({
                'type': 'info' if mom_growth > 0 else 'success',
                'title': f"{'Increase' if mom_growth > 0 else 'Decrease'} in New Issues",
                'message': f"{abs(mom_growth):.0f}% {'more' if mom_growth > 0 else 'fewer'} issues this month vs last month.",
                'metric': f"{trends['current_month_created']} this month"
            })
        # Performance insights
        avg_resolution = performance.get('avg_resolution_hours', 0)
        if avg_resolution > 0:
            if avg_resolution < 24:
                insights.append({
                    'type': 'success',
                    'title': 'Fast Resolution Time',
                    'message': f"Average resolution time is {avg_resolution:.1f} hours - excellent response!",
                    'metric': f"{avg_resolution:.1f}h avg"
                })
            elif avg_resolution > 72:
                insights.append({
                    'type': 'warning',
                    'title': 'Slow Resolution Time',
                    'message': f"Average resolution time is {avg_resolution:.1f} hours. Consider process optimization.",
                    'metric': f"{avg_resolution:.1f}h avg"
                })
        # Assignee workload
        if summary.get('by_assignee'):
            top_assignee = max(summary['by_assignee'].items(), key=lambda x: x[1])
            if top_assignee[1] > summary['total'] * 0.3:  # More than 30% of all issues
                insights.append({
                    'type': 'warning',
                    'title': 'Unbalanced Workload',
                    'message': f"{top_assignee[0]} has {top_assignee[1]} issues ({top_assignee[1]/summary['total']*100:.0f}%). Consider redistributing.",
                    'metric': f"{top_assignee[1]} issues"
                })
        return insights
# ============================================================================
# Background Refresh Worker
# ============================================================================
def background_refresh_worker(service_desk_id: str, queue_id: Optional[str] = None):
    """Background worker to refresh report data."""
    global _refresh_state
    with _refresh_lock:
        _refresh_state['is_running'] = True
        _refresh_state['status'] = 'running'
        _refresh_state['progress'] = 0
        _refresh_state['error'] = None
    try:
        logger.info(f"🔄 Starting background refresh for desk={service_desk_id}, queue={queue_id}")
        # Update progress
        with _refresh_lock:
            _refresh_state['progress'] = 10
            _refresh_state['status'] = 'initializing'
        # Get issues from queue
        from core.api import load_queue_issues
        issues = []
        if queue_id:
            with _refresh_lock:
                _refresh_state['progress'] = 20
                _refresh_state['status'] = 'fetching_issues'
            logger.info(f"📡 Loading queue issues: desk={service_desk_id}, queue={queue_id}")
            try:
                df, error = load_queue_issues(service_desk_id, queue_id)
                with _refresh_lock:
                    _refresh_state['progress'] = 40
                    _refresh_state['status'] = 'processing_data'
                if error:
                    logger.warning(f"⚠️ Error loading issues: {error}")
                    # Continue with empty list instead of failing
                elif df is not None and not df.empty:
                    # Convert DataFrame to list of dicts
                    issues = df.to_dict('records')
                    logger.info(f"✓ Loaded {len(issues)} issues from queue")
                else:
                    logger.info("ℹ️ DataFrame is empty or None")
            except Exception as load_error:
                logger.error(f"❌ Exception loading issues: {load_error}", exc_info=True)
                # Continue with empty list
        else:
            logger.info("ℹ️ No queue specified, generating summary metrics")
            with _refresh_lock:
                _refresh_state['progress'] = 40
                _refresh_state['status'] = 'processing_data'
        with _refresh_lock:
            _refresh_state['progress'] = 50
            _refresh_state['status'] = 'analyzing'
        # Generate metrics
        metrics = MetricsGenerator.analyze_issues(issues)
        with _refresh_lock:
            _refresh_state['progress'] = 70
            _refresh_state['status'] = 'saving'
        # Save to database
        import json
        conn = get_db()
        now = datetime.now()
        # Adaptive cache TTL: 3 hours for large queues (50+ tickets), 1 hour for small queues
        cache_hours = 3 if len(issues) >= 50 else 1
        expires = now + timedelta(hours=cache_hours)
        logger.info(f"💾 Caching metrics with {cache_hours}h TTL ({len(issues)} issues)")
        conn.execute("""
            INSERT INTO reports_cache (report_type, service_desk_id, queue_id, data, generated_at, expires_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, ('metrics', service_desk_id, queue_id, json.dumps(metrics), now.isoformat(), expires.isoformat()))
        conn.commit()
        with _refresh_lock:
            _refresh_state['progress'] = 100
            _refresh_state['status'] = 'completed'
            _refresh_state['last_refresh'] = now.isoformat()
            _refresh_state['next_refresh'] = expires.isoformat()
        logger.info(f"✅ Background refresh completed: {len(issues)} issues analyzed")
    except Exception as e:
        logger.error(f"❌ Background refresh failed: {e}", exc_info=True)
        # Still save empty metrics to cache so frontend doesn't get stuck
        try:
            import json
            conn = get_db()
            now = datetime.now()
            expires = now + timedelta(minutes=5)  # Short cache for errors
            empty_metrics = {
                'summary': {'total': 0, 'open': 0, 'closed': 0},
                'trends': {},
                'performance': {},
                'insights': [{
                    'type': 'warning',
                    'title': 'Data Unavailable',
                    'message': f'Could not load metrics: {str(e)}'
                }],
                'generated_at': now.isoformat(),
                'error': str(e)
            }
            conn.execute("""
                INSERT INTO reports_cache (report_type, service_desk_id, queue_id, data, generated_at, expires_at)
                VALUES (?, ?, ?, ?, ?, ?)
            """, ('metrics', service_desk_id, queue_id or '', json.dumps(empty_metrics), now.isoformat(), expires.isoformat()))
            conn.commit()
            logger.info("💾 Saved empty metrics to cache to prevent frontend hanging")
        except Exception as cache_error:
            logger.error(f"Failed to save error state to cache: {cache_error}")
        with _refresh_lock:
            _refresh_state['progress'] = 100
            _refresh_state['status'] = 'error'
            _refresh_state['error'] = str(e)
    finally:
        with _refresh_lock:
            _refresh_state['is_running'] = False
            logger.info(f"🏁 Background worker finished (status: {_refresh_state['status']})")
# ============================================================================
# API Endpoints
# ============================================================================
@reports_bp.route('/api/reports/metrics', methods=['GET'])
@handle_api_error
@json_response
@log_decorator(logging.INFO)
@require_credentials
def get_metrics():
    """
    Get intelligent metrics for a service desk/queue.
    Uses cached data if available, triggers background refresh if stale.
    Query Parameters:
        - serviceDeskId: Service Desk ID (required)
        - queueId: Queue ID (optional)
        - forceRefresh: Force immediate refresh (default: false)
        - startDate: Filter start date (YYYY-MM-DD, optional)
        - endDate: Filter end date (YYYY-MM-DD, optional)
    """
    service_desk_id = request.args.get('serviceDeskId', '')
    queue_id = request.args.get('queueId', '')
    force_refresh = request.args.get('forceRefresh', 'false').lower() == 'true'
    start_date = request.args.get('startDate', '')
    end_date = request.args.get('endDate', '')
    if not service_desk_id:
        return {'success': False, 'error': 'serviceDeskId is required'}, 400
    import json
    conn = get_db()
    now = datetime.now().isoformat()
    # Check cache
    if not force_refresh:
        row = conn.execute("""
            SELECT data, generated_at, expires_at 
            FROM reports_cache 
            WHERE report_type = ? AND service_desk_id = ? AND (queue_id = ? OR (queue_id IS NULL AND ? = ''))
            AND expires_at > ?
            ORDER BY generated_at DESC 
            LIMIT 1
        """, ('metrics', service_desk_id, queue_id, queue_id, now)).fetchone()
        if row:
            metrics = json.loads(row[0])
            logger.info(f"✅ Returning cached metrics (generated: {row[1]})")
            # Update refresh state to show cached data is ready
            with _refresh_lock:
                _refresh_state['progress'] = 100
                _refresh_state['status'] = 'completed'
                _refresh_state['is_running'] = False
                _refresh_state['last_refresh'] = row[1]
                _refresh_state['next_refresh'] = row[2]
            return {
                'success': True,
                'metrics': metrics,
                'cached': True,
                'generated_at': row[1],
                'expires_at': row[2]
            }
    # No cache or force refresh - trigger background refresh
    if not _refresh_state['is_running']:
        thread = threading.Thread(
            target=background_refresh_worker,
            args=(service_desk_id, queue_id),
            daemon=True
        )
        thread.start()
        logger.info(f"🚀 Triggered background refresh for desk={service_desk_id}, queue={queue_id}")
    # Return current refresh status
    return {
        'success': True,
        'message': 'Metrics refresh in progress',
        'refresh_status': dict(_refresh_state),
        'cached': False
    }, 202  # Accepted
@reports_bp.route('/api/reports/refresh-status', methods=['GET'])
@handle_api_error
@json_response
@log_decorator(logging.INFO)
def get_refresh_status():
    """Get current status of background refresh."""
    return {
        'success': True,
        'status': dict(_refresh_state)
    }
@reports_bp.route('/api/reports/summary', methods=['GET'])
@handle_api_error
@json_response
@log_decorator(logging.INFO)
@require_credentials
def get_reports_summary():
    """Get summary of all available reports."""
    import json
    conn = get_db()
    rows = conn.execute("""
        SELECT report_type, service_desk_id, queue_id, generated_at, expires_at
        FROM reports_cache
        WHERE expires_at > ?
        ORDER BY generated_at DESC
    """, (datetime.now().isoformat(),)).fetchall()
    reports = [{
        'type': row[0],
        'service_desk_id': row[1],
        'queue_id': row[2],
        'generated_at': row[3],
        'expires_at': row[4]
    } for row in rows]
    return {
        'success': True,
        'reports': reports,
        'count': len(reports)
    }
@reports_bp.route('/api/reports/cleanup', methods=['POST'])
@handle_api_error
@json_response
@log_decorator(logging.INFO)
@require_credentials
def cleanup_reports():
    """Clean up expired reports from cache."""
    conn = get_db()
    result = conn.execute("""
        DELETE FROM reports_cache
        WHERE expires_at < ?
    """, (datetime.now().isoformat(),))
    conn.commit()
    deleted = result.rowcount
    logger.info(f"🗑️ Cleaned up {deleted} expired reports")
    return {
        'success': True,
        'deleted': deleted,
        'message': f'Cleaned up {deleted} expired reports'
    }
@reports_bp.route('/api/reports/export/<format>', methods=['GET'])
@handle_api_error
@require_credentials
def export_report(format):
    """
    Export report in specified format (csv, json, excel).
    Query Parameters:
        - serviceDeskId: Service Desk ID
        - queueId: Queue ID (optional)
        - startDate: Start date (YYYY-MM-DD)
        - endDate: End date (YYYY-MM-DD)
    """
    import csv
    import io
    from flask import send_file, make_response
    service_desk_id = request.args.get('serviceDeskId', '')
    queue_id = request.args.get('queueId', '')
    start_date = request.args.get('startDate', '')
    end_date = request.args.get('endDate', '')
    if not service_desk_id:
        return {'success': False, 'error': 'serviceDeskId is required'}, 400
    # Get metrics from cache or generate
    import json as json_lib
    conn = get_db()
    now = datetime.now().isoformat()
    row = conn.execute("""
        SELECT data 
        FROM reports_cache 
        WHERE report_type = ? AND service_desk_id = ? AND (queue_id = ? OR (queue_id IS NULL AND ? = ''))
        AND expires_at > ?
        ORDER BY generated_at DESC 
        LIMIT 1
    """, ('metrics', service_desk_id, queue_id, queue_id, now)).fetchone()
    if not row:
        return {'success': False, 'error': 'No cached metrics available. Generate report first.'}, 404
    metrics = json_lib.loads(row[0])
    summary = metrics.get('summary', {})
    if format == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        # Header
        writer.writerow(['SpeedyFlow Analytics Report'])
        writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'])
        writer.writerow([f'Service Desk: {service_desk_id}'])
        writer.writerow([])
        # Summary
        writer.writerow(['SUMMARY'])
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Tickets', summary.get('total', 0)])
        writer.writerow(['Open Tickets', summary.get('open', 0)])
        writer.writerow(['Closed Tickets', summary.get('closed', 0)])
        writer.writerow([])
        # By Status
        writer.writerow(['BY STATUS'])
        writer.writerow(['Status', 'Count'])
        for status, count in summary.get('by_status', {}).items():
            writer.writerow([status, count])
        writer.writerow([])
        # By Priority
        writer.writerow(['BY PRIORITY'])
        writer.writerow(['Priority', 'Count'])
        for priority, count in summary.get('by_priority', {}).items():
            writer.writerow([priority, count])
        writer.writerow([])
        # By Assignee
        writer.writerow(['BY ASSIGNEE'])
        writer.writerow(['Assignee', 'Count'])
        for assignee, count in summary.get('by_assignee', {}).items():
            writer.writerow([assignee, count])
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'speedyflow_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
    elif format == 'json':
        response = make_response(json_lib.dumps(metrics, indent=2))
        response.headers['Content-Type'] = 'application/json'
        response.headers['Content-Disposition'] = f'attachment; filename=speedyflow_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        return response
    elif format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Analytics Report"
            # Header styling
            header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            # Title
            ws['A1'] = 'SpeedyFlow Analytics Report'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
            ws['A3'] = f'Service Desk: {service_desk_id}'
            # Summary
            row = 5
            ws[f'A{row}'] = 'SUMMARY'
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            ws[f'A{row}'] = 'Total Tickets'
            ws[f'B{row}'] = summary.get('total', 0)
            row += 1
            ws[f'A{row}'] = 'Open Tickets'
            ws[f'B{row}'] = summary.get('open', 0)
            row += 1
            ws[f'A{row}'] = 'Closed Tickets'
            ws[f'B{row}'] = summary.get('closed', 0)
            row += 2
            # By Status
            ws[f'A{row}'] = 'BY STATUS'
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            row += 1
            for status, count in summary.get('by_status', {}).items():
                ws[f'A{row}'] = status
                ws[f'B{row}'] = count
                row += 1
            # Save to bytes
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'speedyflow_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        except ImportError:
            return {'success': False, 'error': 'openpyxl not installed. Use CSV or JSON format.'}, 400
    else:
        return {'success': False, 'error': f'Unsupported format: {format}'}, 400
@reports_bp.route('/api/reports/compare', methods=['GET'])
@handle_api_error
@json_response
@log_decorator(logging.INFO)
@require_credentials
def compare_periods():
    """
    Compare metrics between two periods.
    Query Parameters:
        - serviceDeskId: Service Desk ID
        - period1Start, period1End: First period dates
        - period2Start, period2End: Second period dates
    """
    service_desk_id = request.args.get('serviceDeskId', '')
    # For now, return month-over-month comparison from cached metrics
    import json as json_lib
    conn = get_db()
    now = datetime.now().isoformat()
    row = conn.execute("""
        SELECT data 
        FROM reports_cache 
        WHERE report_type = ? AND service_desk_id = ?
        AND expires_at > ?
        ORDER BY generated_at DESC 
        LIMIT 1
    """, ('metrics', service_desk_id, now)).fetchone()
    if not row:
        return {'success': False, 'error': 'No cached metrics available'}, 404
    metrics = json_lib.loads(row[0])
    trends = metrics.get('trends', {})
    return {
        'success': True,
        'comparison': {
            'current_month': trends.get('current_month_created', 0),
            'last_month': trends.get('last_month_created', 0),
            'growth_percent': trends.get('mom_growth', 0),
            'trend': 'up' if trends.get('mom_growth', 0) > 0 else 'down'
        }
    }
